"""Excel renderer for Phase 14 report generation.

Produces a multi-sheet .xlsx workbook:
  - Executive Summary
  - Findings
  - Charts (severity bar chart)
  - Compliance (if mapping data is present)
"""
from io import BytesIO
from typing import List

from app.core.reporting.base_renderer import BaseRenderer


class ExcelRenderer(BaseRenderer):
    """Renders a ReportV2 + findings list into an openpyxl workbook."""

    def render(self, report, findings: List) -> bytes:  # type: ignore[override]
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        wb = Workbook()

        # ── Sheet 1: Executive Summary ──────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        ws_summary["A1"] = report.title
        ws_summary["A1"].font = Font(size=18, bold=True)
        ws_summary.merge_cells("A1:D1")

        ws_summary["A2"] = "Generated at"
        ws_summary["B2"] = (
            report.generated_at.isoformat() if report.generated_at else ""
        )
        ws_summary["A3"] = "Project ID"
        ws_summary["B3"] = report.project_id
        ws_summary["A4"] = "Report Type"
        ws_summary["B4"] = str(report.report_type)
        ws_summary["A5"] = "Status"
        ws_summary["B5"] = str(report.status)

        row = 7
        ws_summary[f"A{row}"] = "Summary"
        ws_summary[f"A{row}"].font = Font(bold=True)
        row += 1

        summary = report.summary_metadata or {}
        for key, value in summary.items():
            if key == "by_tool":
                continue
            ws_summary[f"A{row}"] = key.replace("_", " ").title()
            ws_summary[f"B{row}"] = str(value)
            row += 1

        # ── Sheet 2: Findings ───────────────────────────────────────────────
        ws_findings = wb.create_sheet("Findings")
        headers = ["Title", "Severity", "Host", "Tool", "Risk Score", "Description", "Remediation"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws_findings.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        severity_colors = {
            "critical": "DC3545",
            "high": "FD7E14",
            "medium": "FFC107",
            "low": "28A745",
            "info": "17A2B8",
        }

        for row_idx, f in enumerate(findings, 2):
            sev = f.severity.value if hasattr(f.severity, "value") else str(f.severity)
            ws_findings.cell(row=row_idx, column=1, value=f.title)
            sev_cell = ws_findings.cell(row=row_idx, column=2, value=sev.upper())
            color = severity_colors.get(sev.lower(), "AAAAAA")
            sev_cell.fill = PatternFill("solid", fgColor=color)
            sev_cell.font = Font(bold=True, color="FFFFFF")
            ws_findings.cell(row=row_idx, column=3, value=f.host or "")
            ws_findings.cell(row=row_idx, column=4, value=f.tool_name or "")
            ws_findings.cell(row=row_idx, column=5, value=round(f.risk_score or 0.0, 2))
            ws_findings.cell(row=row_idx, column=6, value=(f.description or "")[:200])
            ws_findings.cell(row=row_idx, column=7, value=(getattr(f, "remediation", "") or "")[:200])

        # Auto-width approximation
        for col in ws_findings.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws_findings.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # ── Sheet 3: Charts ─────────────────────────────────────────────────
        ws_charts = wb.create_sheet("Charts")
        ws_charts["A1"] = "Severity"
        ws_charts["B1"] = "Count"
        chart_data = [
            ("Critical", summary.get("critical", 0)),
            ("High",     summary.get("high", 0)),
            ("Medium",   summary.get("medium", 0)),
            ("Low",      summary.get("low", 0)),
            ("Info",     summary.get("info", 0)),
        ]
        for i, (label, count) in enumerate(chart_data, 2):
            ws_charts[f"A{i}"] = label
            ws_charts[f"B{i}"] = count

        chart = BarChart()
        chart.type = "col"
        chart.title = "Findings by Severity"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Severity"
        data_ref = Reference(ws_charts, min_col=2, min_row=1, max_row=6)
        cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=6)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_charts.add_chart(chart, "D2")

        # ── Sheet 4: Compliance (optional) ──────────────────────────────────
        if report.compliance_mapping:
            ws_comp = wb.create_sheet("Compliance")
            ws_comp["A1"] = "Requirement ID"
            ws_comp["B1"] = "Status"
            ws_comp["C1"] = "Evidence"
            for i, item in enumerate(report.compliance_mapping.audit_findings or [], 2):
                ws_comp.cell(row=i, column=1, value=item.get("requirement_id", ""))
                ws_comp.cell(row=i, column=2, value=item.get("status", ""))
                ws_comp.cell(row=i, column=3, value=str(item.get("evidence", ""))[:200])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def extension(self) -> str:
        return "xlsx"
